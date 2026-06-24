#!/usr/bin/env python3
"""plexcel — CSV to styled XLSX using YAML config."""
import csv
import re
import sys
from datetime import datetime
from pathlib import Path

import yaml
import openpyxl
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.datavalidation import DataValidation

# --- Color helpers ---

NAMED_COLORS = {
    'white': 'FFFFFFFF', 'black': 'FF000000', 'red': 'FFFF0000',
    'green': 'FF92D050', 'blue': 'FF1F3864', 'yellow': 'FFFFFF00',
    'orange': 'FFFF8C00', 'purple': 'FF7030A0', 'gray': 'FF808080',
}

def to_argb(color):
    """Convert color name or #hex to 8-char ARGB."""
    if not color:
        return None
    color = color.strip().lower()
    if color in NAMED_COLORS:
        return NAMED_COLORS[color]
    color = color.lstrip('#')
    if len(color) == 6:
        return 'FF' + color.upper()
    return color.upper()

# --- Shorthand parsers ---

def parse_font(shorthand):
    """Parse 'Arial 14 bold italic white' into Font kwargs."""
    if isinstance(shorthand, dict):
        return shorthand
    parts = shorthand.split()
    kwargs = {'name': 'Arial', 'size': 11, 'bold': False, 'italic': False, 'color': 'FF000000'}
    i = 0
    # Font family (non-numeric, non-keyword first token)
    if parts and not parts[0].isdigit() and parts[0] not in ('bold', 'italic') and not parts[0].startswith('#'):
        kwargs['name'] = parts[0]
        i = 1
    # Size
    if i < len(parts) and parts[i].isdigit():
        kwargs['size'] = int(parts[i])
        i += 1
    # Modifiers and color
    while i < len(parts):
        p = parts[i].lower()
        if p == 'bold':
            kwargs['bold'] = True
        elif p == 'italic':
            kwargs['italic'] = True
        else:
            kwargs['color'] = to_argb(p)
        i += 1
    return kwargs


def parse_fill(shorthand):
    """Parse '#1F3864' or color name into fill color ARGB."""
    if isinstance(shorthand, dict):
        return shorthand.get('color', shorthand.get('fgColor'))
    return to_argb(str(shorthand))


def parse_align(shorthand):
    """Parse 'center center wrap' into Alignment kwargs."""
    if isinstance(shorthand, dict):
        return shorthand
    parts = shorthand.split()
    kwargs = {}
    h_vals = ('left', 'center', 'right', 'justify')
    v_vals = ('top', 'center', 'bottom')
    for p in parts:
        p = p.lower()
        if p in h_vals and 'horizontal' not in kwargs:
            kwargs['horizontal'] = p
        elif p in v_vals and 'horizontal' in kwargs:
            kwargs['vertical'] = p
        elif p == 'wrap':
            kwargs['wrap_text'] = True
    return kwargs


def parse_border(shorthand):
    """Parse border shorthand into dict of {edge: Side}."""
    if isinstance(shorthand, str):
        # Single style applies to all edges
        side = Side(style=shorthand)
        return {'left': side, 'right': side, 'top': side, 'bottom': side}
    sides = {}
    for edge, style in shorthand.items():
        if isinstance(style, str):
            sides[edge] = Side(style=style)
        else:
            sides[edge] = Side(style=style.get('style', 'thin'), color=to_argb(style.get('color', 'black')))
    return sides

# --- Style application ---

def build_style(style_def):
    """Convert a YAML style definition into openpyxl style objects."""
    result = {}
    if 'font' in style_def:
        fk = parse_font(style_def['font'])
        result['font'] = Font(name=fk['name'], size=fk['size'], bold=fk['bold'],
                              italic=fk['italic'], color=fk.get('color', 'FF000000'))
    if 'fill' in style_def:
        color = parse_fill(style_def['fill'])
        result['fill'] = PatternFill(patternType='solid', fgColor=color)
    if 'align' in style_def:
        ak = parse_align(style_def['align'])
        result['alignment'] = Alignment(**ak)
    if 'border' in style_def:
        result['border'] = Border(**parse_border(style_def['border']))
    return result


def apply_built_style(cell, built):
    """Apply pre-built style objects to a cell."""
    if 'font' in built:
        cell.font = built['font']
    if 'fill' in built:
        cell.fill = built['fill']
    if 'alignment' in built:
        cell.alignment = built['alignment']
    if 'border' in built:
        cell.border = built['border']

# --- Value coercion ---

def coerce_value(value):
    if not isinstance(value, str) or not value.strip():
        return value
    v = value.strip()
    if v.lower() in ('true', 'false'):
        return v.lower() == 'true'
    try:
        if '.' not in v and 'e' not in v.lower():
            return int(v)
    except ValueError:
        pass
    try:
        return float(v)
    except ValueError:
        pass
    for fmt in ('%Y-%m-%d', '%m/%d/%Y'):
        try:
            return datetime.strptime(v, fmt)
        except ValueError:
            continue
    return value.replace('\\n', '\n')

# --- Column auto-fit ---

def auto_fit_columns(ws, wrap_cols, max_width=55, min_width=8):
    for row in ws.iter_rows():
        for cell in row:
            col_letter = get_column_letter(cell.column)
            if col_letter in wrap_cols:
                continue
            val = str(cell.value) if cell.value is not None else ''
            if isinstance(cell.value, datetime):
                val = '2026-07-01'
            cur = ws.column_dimensions[col_letter].width or 0
            ws.column_dimensions[col_letter].width = max(cur, min(len(val) + 3, max_width), min_width)
    for col_letter in wrap_cols:
        ws.column_dimensions[col_letter].width = max_width

# --- Main sheet builder ---

def process_sheet(ws, csv_file, sheet_cfg, built_styles, defaults):
    if not Path(csv_file).exists():
        print(f"Error: CSV file not found: {csv_file}")
        sys.exit(1)

    wrap_cols = set(sheet_cfg.get('wrap_columns', []))
    date_fmt = defaults.get('date_format', 'yyyy-mm-dd')
    default_align = Alignment(**parse_align(defaults.get('align', 'left center')))
    wrap_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # Load CSV
    with open(csv_file, 'r', encoding='utf-8') as f:
        for row_idx, row in enumerate(csv.reader(f), 1):
            for col_idx, value in enumerate(row, 1):
                coerced = coerce_value(value)
                cell = ws.cell(row=row_idx, column=col_idx, value=coerced)
                col_letter = get_column_letter(col_idx)
                if row_idx == 1 and 'header' in built_styles:
                    apply_built_style(cell, built_styles['header'])
                elif col_letter in wrap_cols:
                    cell.alignment = wrap_align
                else:
                    cell.alignment = default_align
                if isinstance(coerced, datetime):
                    cell.number_format = date_fmt

    # Auto-fit columns
    auto_fit_columns(ws, wrap_cols)

    # Explicit column width overrides
    for col_letter, width in sheet_cfg.get('column_widths', {}).items():
        ws.column_dimensions[col_letter].width = width

    # Row heights
    row_height = defaults.get('row_height', 25.5)
    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].height = row_height
    for row_str, height in sheet_cfg.get('row_heights', {}).items():
        ws.row_dimensions[int(row_str)].height = height

    # Auto-merge row 1 if only first cell has content
    if ws.max_column > 1:
        row1 = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        if sum(1 for v in row1 if v) == 1:
            ws.merge_cells(f"A1:{get_column_letter(ws.max_column)}1")

    # Explicit merges (supports both 'merge' and 'merged_cells' keys)
    for rng in sheet_cfg.get('merge', sheet_cfg.get('merged_cells', [])):
        ws.merge_cells(rng)

    # Freeze panes — supports freeze_pane: "A2" or freeze_row: 1
    freeze_pane = sheet_cfg.get('freeze_pane')
    freeze_row = sheet_cfg.get('freeze_row', defaults.get('freeze_row'))
    if freeze_pane:
        ws.freeze_panes = freeze_pane
    elif freeze_row:
        ws.freeze_panes = f"A{freeze_row + 1}"

    # Auto-filter
    af = sheet_cfg.get('auto_filter', defaults.get('auto_filter'))
    if af:
        if af is True:
            ws.auto_filter.ref = ws.dimensions
        elif isinstance(af, str) and af.startswith('row'):
            r = int(af.split()[1])
            ws.auto_filter.ref = f"A{r}:{get_column_letter(ws.max_column)}{r}"
        else:
            ws.auto_filter.ref = af

    # Cell styles (supports dict {"range": "style"} or list [{range, style}])
    cell_styles = sheet_cfg.get('cell_styles', {})
    if isinstance(cell_styles, list):
        cell_styles = {item['range']: item['style'] for item in cell_styles}
    for rng, style_name in cell_styles.items():
        if style_name not in built_styles:
            print(f"Warning: Style '{style_name}' not found, skipping")
            continue
        cells = ws[rng]
        if isinstance(cells, openpyxl.cell.Cell):
            apply_built_style(cells, built_styles[style_name])
        else:
            for row_cells in cells:
                if not isinstance(row_cells, tuple):
                    row_cells = (row_cells,)
                for c in row_cells:
                    apply_built_style(c, built_styles[style_name])

    # Row styles (match by cell value)
    for rule in sheet_cfg.get('row_styles', []):
        match = rule['match']
        style_name = rule['style']
        if style_name not in built_styles:
            continue
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for col_letter, expected in match.items():
                col_idx = column_index_from_string(col_letter)
                cell_val = row[col_idx - 1].value
                if str(cell_val).strip() == str(expected).strip():
                    for c in row:
                        apply_built_style(c, built_styles[style_name])

    # Dropdowns (supports dict {col: [values]} or list [{range, values, ...}])
    dropdowns = sheet_cfg.get('dropdowns', {})
    if isinstance(dropdowns, dict):
        for col_letter, values in dropdowns.items():
            val_str = ','.join(str(v) for v in values)
            dv = DataValidation(type="list", formula1=f'"{val_str}"', allow_blank=True)
            start_row = (sheet_cfg.get('freeze_row', 1)) + 1
            dv.add(f"{col_letter}{start_row}:{col_letter}1000")
            ws.add_data_validation(dv)
    elif isinstance(dropdowns, list):
        for dd in dropdowns:
            val_str = dd['values'] if isinstance(dd['values'], str) else ','.join(str(v) for v in dd['values'])
            dv = DataValidation(type="list", formula1=f'"{val_str}"',
                                allow_blank=dd.get('allow_blank', True))
            if 'promptTitle' in dd:
                dv.promptTitle = dd['promptTitle']
            if 'prompt' in dd:
                dv.prompt = dd['prompt']
            if 'errorTitle' in dd:
                dv.errorTitle = dd['errorTitle']
            if 'error' in dd:
                dv.error = dd['error']
            dv.add(dd['range'])
            ws.add_data_validation(dv)

    # Highlight (legacy shorthand conditional formatting)
    for rule in sheet_cfg.get('highlight', []):
        when = rule['when']
        for col_letter, expected in when.items():
            start_row = (sheet_cfg.get('freeze_row', 1)) + 1
            rng = f"A{start_row}:{get_column_letter(ws.max_column)}{ws.max_row}"
            formula = f'${col_letter}{start_row}="{expected}"'
            fill = PatternFill(bgColor=to_argb(rule.get('fill', ''))) if 'fill' in rule else None
            font = Font(color=to_argb(rule.get('font', ''))) if 'font' in rule else None
            ws.conditional_formatting.add(rng, FormulaRule(formula=[formula], fill=fill, font=font))

    # Conditional formatting (full)
    for rule in sheet_cfg.get('conditional_formatting', []):
        rng = rule['range']
        style = rule.get('style', {})
        fill = PatternFill(bgColor=to_argb(style['fill_color'])) if 'fill_color' in style else None
        font = Font(color=to_argb(style['font_color'])) if 'font_color' in style else None
        if rule['type'] == 'cell_is':
            ws.conditional_formatting.add(rng, CellIsRule(
                operator=rule['operator'], formula=rule['formula'], fill=fill, font=font))
        elif rule['type'] == 'formula':
            ws.conditional_formatting.add(rng, FormulaRule(
                formula=rule['formula'], fill=fill, font=font))

    # Number formats
    for nf in sheet_cfg.get('number_formats', []):
        fmt = nf['format']
        if 'range' in nf:
            for row_cells in ws[nf['range']]:
                if not isinstance(row_cells, tuple):
                    row_cells = (row_cells,)
                for c in row_cells:
                    c.number_format = fmt
        elif 'column' in nf:
            col = nf['column']
            start = nf.get('start_row', 2)
            end = nf.get('end_row', ws.max_row)
            for r in range(start, end + 1):
                ws[f"{col}{r}"].number_format = fmt

    # Formulas
    for f_cfg in sheet_cfg.get('formulas', []):
        if 'cell' in f_cfg:
            ws[f_cfg['cell']].value = f_cfg['formula']
        elif 'column' in f_cfg:
            for row in range(f_cfg['start_row'], f_cfg.get('end_row', ws.max_row) + 1):
                ws[f"{f_cfg['column']}{row}"].value = f_cfg['formula'].replace('{row}', str(row))


def main():
    if len(sys.argv) < 2:
        print("Usage: python csv_to_xlsx.py output.xlsx [workbook.yaml]")
        print("       python csv_to_xlsx.py output.xlsx file.csv:Sheet1 [file.csv:Sheet2 ...]")
        sys.exit(1)

    output_file = sys.argv[1]

    # Load shared styles — check current dir, then script's directory
    styles_path = Path('styles.yaml')
    if not styles_path.exists():
        styles_path = Path(__file__).parent / 'styles.yaml'
    if not styles_path.exists():
        print(f"Error: No styles.yaml found in current dir or {Path(__file__).parent}")
        sys.exit(1)

    try:
        with open(styles_path) as f:
            styles_cfg = yaml.safe_load(f)
    except Exception as e:
        print(f"Error reading styles.yaml: {e}")
        sys.exit(1)

    defaults = styles_cfg.get('defaults', {})
    built_styles = {name: build_style(sdef) for name, sdef in styles_cfg.get('styles', {}).items()}

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    try:
        # YAML config mode
        if len(sys.argv) == 3 and sys.argv[2].endswith('.yaml'):
            wb_path = Path(sys.argv[2])
            if not wb_path.exists():
                print(f"Error: {sys.argv[2]} not found")
                sys.exit(1)
            with open(wb_path) as f:
                wb_cfg = yaml.safe_load(f)
            for sheet_cfg in wb_cfg['sheets']:
                ws = wb.create_sheet(sheet_cfg.get('name', Path(sheet_cfg['csv']).stem))
                process_sheet(ws, sheet_cfg['csv'], sheet_cfg, built_styles, defaults)

        # Simple mode: csv:SheetName pairs
        else:
            for arg in sys.argv[2:]:
                if ':' not in arg:
                    print(f"Error: Use format file.csv:SheetName, got '{arg}'")
                    sys.exit(1)
                csv_file, sheet_name = arg.split(':', 1)
                ws = wb.create_sheet(sheet_name)
                process_sheet(ws, csv_file, {}, built_styles, defaults)

        wb.save(output_file)
        print(f"Created {output_file} with {len(wb.sheetnames)} sheets")

    except PermissionError:
        print(f"Error: Cannot write to {output_file}. Is it open in Excel?")
        sys.exit(1)
    except Exception as e:
        print(f"Error: {e}")
        sys.exit(1)


if __name__ == '__main__':
    main()
